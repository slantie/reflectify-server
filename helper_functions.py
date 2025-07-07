import re
import openpyxl
import numpy as np
import pandas as pd
from typing import Dict, List, Set, Any, Optional

def clean_excel_cell(value: Any) -> Optional[Any]:
    """Cleans Excel cell values by stripping whitespace and converting empty strings to None."""
    if value is None:
        return None
    if isinstance(value, str):
        cleaned_str = value.strip()
        return cleaned_str if cleaned_str != '' else None
    return value

def extract_sheet_data(workbook: openpyxl.workbook.workbook.Workbook, sheet_name: str) -> pd.DataFrame:
    """Extracts tabular data from Excel sheet, handling merged cells and dynamic header detection."""
    sheet = workbook[sheet_name]
    merged_ranges = sheet.merged_cells.ranges
    
    # Create new workbook to handle unmerged cells
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Copy and clean all cell data
    for row in sheet.rows:
        for cell in row:
            cleaned_value = clean_excel_cell(cell.value)
            new_sheet.cell(row=cell.row, column=cell.column, value=cleaned_value)
    
    # Propagate values for merged cells
    for merged_range in merged_ranges:
        value_to_propagate = clean_excel_cell(sheet.cell(merged_range.min_row, merged_range.min_col).value)
        
        # Fill all cells within merged range
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for col_index in range(merged_range.min_col, merged_range.max_col + 1):
                new_sheet.cell(row=row_index, column=col_index, value=value_to_propagate)
    
    # Convert to DataFrame
    data_frame = pd.DataFrame(new_sheet.values)

    # Detect header row by finding 'SLOT' in second column
    try:
        header_index = data_frame[data_frame.iloc[:, 1].astype(str) == 'SLOT'].index[0]
    except IndexError:
        return pd.DataFrame()
        
    header_row_values = data_frame.iloc[header_index]
    
    # Clean the header row values
    for idx, name in enumerate(header_row_values):
        if isinstance(name, str) and re.match(r'^L\d+$', name):
            # Remove the column as well
            data_frame.drop(columns=[name], inplace=True, errors='ignore')
            header_row_values[idx] = None
            continue

        if isinstance(name, str) and "/" in name:
            name_parts = name.split('/')
            header_row_values[idx] = name_parts[0].replace(' ', '').strip()
    
    # Trim columns based on header row NaN values
    first_unwanted_col_index = -1
    for i, col_value in enumerate(header_row_values):
        if pd.isna(col_value):
            first_unwanted_col_index = i
            break
            
    # Remove unwanted columns if found
    if first_unwanted_col_index != -1:
        data_frame = data_frame.iloc[:, :first_unwanted_col_index]
    
    # Set column names and clean data
    data_frame.columns = header_row_values.iloc[:data_frame.shape[1]]
    data_frame = data_frame.iloc[header_index + 1:].reset_index(drop=True)
    data_frame.replace('', np.nan, inplace=True)
    data_frame.dropna(how='all', inplace=True)
    
    return data_frame

def extract_subject_details(subject_string: str) -> Optional[Dict[str, Any]]:
    if not isinstance(subject_string, str) or not subject_string.strip() or 'TUT' in subject_string.upper():
        return None

    # Remove parenthesis content like "(HJP)", "(TNG)", etc.
    subject_string = re.sub(r'\(.*?\)', '', subject_string).strip()

    # Tokenize by whitespace
    tokens = subject_string.strip().split()
    if not tokens:
        return {"error": f"Empty subject string: {subject_string}"}

    subject_code = None
    semester = None
    division_info = []

    # Step 1: Identify semester (first integer found)
    for token in tokens:
        if re.match(r'^\d+$', token):  # pure integer
            semester = int(token)
            continue

    if not semester:
        # Try matching like "7A1/B2"
        for token in tokens:
            sem_match = re.match(r'^(\d+)[A-Z]', token, re.IGNORECASE)
            if sem_match:
                semester = int(sem_match.group(1))
                break

    if not semester:
        return {"error": f"Semester not found in: {subject_string}"}

    # Step 2: Identify subject code (all alphabets, usually 2-4 letters)
    for token in tokens:
        if re.match(r'^[A-Z]{2,5}$', token, re.IGNORECASE):
            subject_code = token.strip().upper()
            break

    if not subject_code:
        return {"error": f"Subject code not found in: {subject_string}"}

    # Step 3: Extract remaining segment for division and batch
    # Remove semester and subject_code from tokens
    remaining_tokens = [t for t in tokens if t not in [str(semester), subject_code]]
    division_string = ''.join(remaining_tokens).replace(' ', '')  # merge like "A1/B2"

    division_batches: List[Dict[str, Optional[str]]] = []

    if 'ALL' in division_string.upper():
        division_batches = [{'division': 'ALL', 'batch': None}]
    else:
        division_segments = division_string.split('/')

        for segment in division_segments:
            division_letter = ''.join(char for char in segment if char.isalpha()).strip()

            batch_value: Optional[str] = None
            batch_match = re.search(r'(\d+\*?)$', segment)
            if batch_match:
                batch_value = batch_match.group(1)

            if division_letter:
                division_batches.append({'division': division_letter, 'batch': batch_value})

    # Remove duplicates
    if not division_batches or division_batches[0].get('division') != 'ALL':
        unique_division_batches: List[Dict[str, Optional[str]]] = []
        seen_tuples: Set[tuple] = set()
        for item in division_batches:
            item_tuple = (item['division'], item['batch'])
            if item_tuple not in seen_tuples:
                unique_division_batches.append(item)
                seen_tuples.add(item_tuple)

        division_batches = sorted(unique_division_batches, key=lambda x: (x['division'], x['batch'] if x['batch'] is not None else ''))

    return {
        'subject_code': subject_code,
        'semester': semester,
        'division_batches': division_batches
    }

def build_faculty_schedules(processed_data_frame: pd.DataFrame) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """Creates a dictionary mapping each faculty member to their weekly schedule."""
    
    DAY_MAPPING: Dict[str, str] = {
        'MON': 'Monday', 'MONDAY': 'Monday',
        'TUES': 'Tuesday', 'TUESDAY': 'Tuesday',
        'WED': 'Wednesday', 'WEDNESDAY': 'Wednesday',
        'THUR': 'Thursday', 'THURSDAY': 'Thursday',
        'FRI': 'Friday', 'FRIDAY': 'Friday',
        'SAT': 'Saturday', 'SATURDAY': 'Saturday'
    }

    faculty_master: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}

    # Identify faculty names from DataFrame columns
    faculty_names = [
        col for col in processed_data_frame.columns[2:]
    ]
    
    # Initialize empty schedule for each faculty
    for faculty in faculty_names:
        faculty_master[faculty] = {
            'Monday': [], 'Tuesday': [], 'Wednesday': [],
            'Thursday': [], 'Friday': [], 'Saturday': []
        }

    # Process schedules for each faculty
    for faculty in faculty_names:
        faculty_column = processed_data_frame[faculty]

        row_index = 0
        while row_index < len(faculty_column):
            raw_day_value = processed_data_frame.iloc[row_index, 0]
            raw_time_slot_start = processed_data_frame.iloc[row_index, 1]
            raw_subject_string = faculty_column.iloc[row_index]

            # Process day value
            day_string = str(raw_day_value).strip().upper()
            day = DAY_MAPPING.get(day_string)

            # Check if cell contains valid data
            if pd.notna(raw_subject_string) and raw_subject_string != '' and day is not None:
                # Convert time slot to int
                try:
                    time_slot_start = int(raw_time_slot_start)
                except (ValueError, TypeError):
                    row_index += 1
                    continue

                current_subject_value = raw_subject_string
                block_start_row_index = row_index

                # Find end of contiguous block for current subject
                block_end_row_index = row_index
                while (block_end_row_index + 1 < len(faculty_column) and
                       faculty_column.iloc[block_end_row_index + 1] == current_subject_value and
                       str(processed_data_frame.iloc[block_end_row_index + 1, 0]).strip().upper() == day_string):
                    block_end_row_index += 1

                # Calculate block length
                block_length = (block_end_row_index - block_start_row_index) + 1

                # Determine activity type based on block length
                if block_length == 2:
                    activity_type = 'Lab'
                elif block_length == 1:
                    activity_type = 'Lecture'
                else:
                    activity_type = 'Unknown'

                # Create schedule entry
                schedule_entry = {
                    'subject_string': current_subject_value,
                    'type': activity_type,
                    'time_slot': time_slot_start,
                    'parsed_subject_info': extract_subject_details(current_subject_value)
                }

                faculty_master[faculty][day].append(schedule_entry)

                # Advance to end of processed block
                row_index = block_end_row_index + 1
            else:
                # Move to next row if cell is empty/invalid
                row_index += 1

    return faculty_master

def standardize_time_slots(data_frame: pd.DataFrame) -> pd.DataFrame:
    """
    Formats the 'Time_Slot' column for 'Lab' entries in a DataFrame.
    
    For 'Lab' entries, it converts the single starting time slot (e.g., 3)
    into a two-slot range string (e.g., "3-4"). Lecture time slots remain as is.
    The DataFrame is then sorted.

    Args:
        data_frame (pd.DataFrame): The DataFrame containing schedule entries.
                           Expected columns include 'Type' and 'Time_Slot'.

    Returns:
        pd.DataFrame: The DataFrame with 'Time_Slot' formatted for labs,
                      and sorted by 'Day', 'Time_Slot', and 'Batch'.
    """
    df_copy = data_frame.copy() 

    # Explicitly cast 'Time_Slot' to object dtype to allow mixed types (int and str)
    # before applying string formatting. This resolves the FutureWarning.
    df_copy['Time_Slot'] = df_copy['Time_Slot'].astype(object)

    lab_mask = df_copy['Type'] == 'Lab'
    df_copy.loc[lab_mask, 'Time_Slot'] = df_copy.loc[lab_mask, 'Time_Slot'].apply(
        lambda x: f"{int(x)}-{int(x)+1}" if pd.notna(x) else x
    )
    
    def get_time_slot_sort_value(slot: Any) -> Any:
        """Extracts sortable value from time slot."""
        if isinstance(slot, str) and '-' in slot:
            try:
                return int(slot.split('-')[0])
            except ValueError:
                return float('inf') 
        elif pd.isna(slot):
            return float('inf')
        try:
            return int(slot)
        except (ValueError, TypeError):
            return float('inf')

    # Sort DataFrame with custom time slot sorting
    return df_copy.sort_values(
        by=['Day', 'Time_Slot', 'Batch'],
        key=lambda col: col.apply(get_time_slot_sort_value) if col.name == 'Time_Slot' else col.astype(str),
        ignore_index=True
    )

def generate_class_schedules(faculty_schedules: Dict[str, Dict[str, List[Dict[str, Any]]]], parse_errors: Optional[List[str]] = None) -> Dict[str, pd.DataFrame]:
    division_tables: Dict[str, List[Dict[str, Any]]] = {}
    semester_divisions: Dict[int, Set[str]] = {}

    # Ensure parse_errors is a list if provided, or initialize it
    if parse_errors is None:
        parse_errors = []

    # First pass: collect all unique divisions per semester
    for faculty_name, faculty_schedule_by_day in faculty_schedules.items():
        for day, sessions in faculty_schedule_by_day.items():
            for session_entry in sessions:
                parsed_subject_info = session_entry.get('parsed_subject_info')

                if (
                    not isinstance(parsed_subject_info, dict)
                    or 'semester' not in parsed_subject_info
                    or parsed_subject_info['semester'] is None
                ):
                    if isinstance(parsed_subject_info, dict) and 'error' in parsed_subject_info:
                        # Log the error and add to the list
                        error_message = f"Error parsing subject info for faculty {faculty_name}, day {day}: {parsed_subject_info['error']}"
                        print(f"[Parse Error] {error_message}") # Keep the logging intact
                        parse_errors.append(error_message)
                    continue

                semester = parsed_subject_info['semester']

                if semester not in semester_divisions:
                    semester_divisions[semester] = set()

                for db_entry in parsed_subject_info.get('division_batches', []):
                    division_letter = db_entry.get('division')
                    if division_letter and division_letter != 'ALL':
                        semester_divisions[semester].add(division_letter)

    # Second pass: populate division tables
    for faculty_name, faculty_schedule_by_day in faculty_schedules.items():
        for day, sessions in faculty_schedule_by_day.items():
            for session_entry in sessions:
                parsed_subject_info = session_entry.get('parsed_subject_info')

                if (
                    not isinstance(parsed_subject_info, dict)
                    or 'semester' not in parsed_subject_info
                    or parsed_subject_info['semester'] is None
                ):
                    continue

                semester = parsed_subject_info['semester']
                target_division_batch_entries: List[Dict[str, Optional[str]]] = []

                session_division_batches = parsed_subject_info.get('division_batches', [])

                if session_division_batches and session_division_batches[0].get('division') == 'ALL':
                    if semester in semester_divisions:
                        for div in sorted(list(semester_divisions[semester])):
                            target_division_batch_entries.append({'division': div, 'batch': None})
                else:
                    target_division_batch_entries = session_division_batches

                for db_entry in target_division_batch_entries:
                    division = db_entry.get('division')
                    batch = db_entry.get('batch')

                    if division:
                        division_key = f"{semester}{division}"

                        if division_key not in division_tables:
                            division_tables[division_key] = []

                        entry = {
                            'Subject': parsed_subject_info.get('subject_code'),
                            'Type': session_entry.get('type'),
                            'Batch': batch if batch is not None else '-',
                            'Day': day,
                            'Time_Slot': session_entry.get('time_slot'),
                            'Faculty': faculty_name
                        }

                        division_tables[division_key].append(entry)

    # Convert to DataFrames and apply formatting
    final_division_dataframes: Dict[str, pd.DataFrame] = {}
    for division_key, entries_list in division_tables.items():
        if entries_list:
            data_frame = pd.DataFrame(entries_list)
            final_division_dataframes[division_key] = standardize_time_slots(data_frame)
        else:
            final_division_dataframes[division_key] = pd.DataFrame(columns=['Subject', 'Type', 'Batch', 'Day', 'Time_Slot', 'Faculty'])

    return final_division_dataframes

def process_all_timetables(matrix_file_path: str, errors: Optional[List[str]] = None) -> Dict[str, Dict[str, List[Dict[str, Any]]]]:
    """Loads Excel workbook, processes each sheet, and consolidates faculty schedules."""
    if errors is None:
        errors = []
    try:
        workbook = openpyxl.load_workbook(matrix_file_path)
    except FileNotFoundError:
        errors.append(f"Error: Timetable file not found at {matrix_file_path}")
        return {}
    except Exception as e:
        errors.append(f"Error loading Excel workbook: {str(e)}")
        return {}

    all_faculty_schedules: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        processed_data_frame = extract_sheet_data(workbook, sheet_name)
        
        if processed_data_frame.empty:
            errors.append(f"Warning: No valid data extracted from sheet '{sheet_name}'. It might be empty or malformed.")
            continue

        sheet_schedules = build_faculty_schedules(
            processed_data_frame
        )
        
        # Merge schedules from current sheet into consolidated schedules
        for faculty_name, schedule_by_day in sheet_schedules.items():
            if faculty_name not in all_faculty_schedules:
                all_faculty_schedules[faculty_name] = schedule_by_day
            else:
                for day_name in schedule_by_day:
                    all_faculty_schedules[faculty_name][day_name].extend(schedule_by_day[day_name])

    # Sort schedule entries by time slot for each faculty and day
    for faculty_name, schedule_by_day in all_faculty_schedules.items():
        for day_name, entries in schedule_by_day.items():
            if entries and 'time_slot' in entries[0]:
                all_faculty_schedules[faculty_name][day_name] = sorted(entries, key=lambda x: x['time_slot'])
            
    return all_faculty_schedules

def get_division_course_catalog(division_tables: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    """Creates condensed timetable DataFrames for each class division."""
    condensed_tables: Dict[str, pd.DataFrame] = {}
    
    # Iterate through each division's detailed timetable DataFrame
    for division_key, data_frame in division_tables.items():
        # Drop time-related columns
        condensed_df = data_frame.drop(['Time_Slot', 'Day'], axis=1, errors='ignore')
        
        # Remove duplicate entries
        condensed_df = condensed_df.drop_duplicates()
        
        # Sort for better readability
        condensed_df = condensed_df.sort_values(
            by=['Subject', 'Batch'],
            key=lambda col: col.astype(str) if col.name == 'Batch' else col,
            ignore_index=True
        )
        
        condensed_tables[division_key] = condensed_df
    
    return condensed_tables

def build_hierarchical_schedule(condensed_division_tables: Dict[str, pd.DataFrame], department: str, college: str = "LDRP-ITR") -> Dict[str, Any]:
    """Creates a final hierarchical dictionary consolidating all timetable information grouped by college, department, semester, division, and subject."""
    # Initialize the main dictionary structure with college and department
    final_consolidated_data: Dict[str, Any] = {
        college: {
            department: {}
        }
    }
    
    # Iterate through each condensed division timetable
    for division_key, data_frame in condensed_division_tables.items():
        # Extract semester and division from division_key (e.g., "5A" -> semester "5", division "A")
        semester_str = division_key[0]
        division_str = division_key[1:]
        
        # Initialize semester entry if it doesn't exist
        if semester_str not in final_consolidated_data[college][department]:
            final_consolidated_data[college][department][semester_str] = {}
            
        # Initialize division entry if it doesn't exist
        if division_str not in final_consolidated_data[college][department][semester_str]:
            final_consolidated_data[college][department][semester_str][division_str] = {}
            
        # Process each unique subject within the current division's DataFrame
        for subject_code, subject_data_group in data_frame.groupby('Subject'):
            subject_details: Dict[str, Any] = {
                'lectures': {},
                'labs': {}
            }
            
            # Process lectures for the current subject
            lectures_df = subject_data_group[subject_data_group['Type'] == 'Lecture']
            if not lectures_df.empty:
                subject_details['lectures'] = {
                    'designated_faculty': lectures_df['Faculty'].iloc[0]
                }
            
            # Process labs for the current subject
            labs_df = subject_data_group[subject_data_group['Type'] == 'Lab']
            if not labs_df.empty:
                # Associate each batch with its designated faculty
                subject_details['labs'] = {
                    str(batch): {'designated_faculty': faculty}
                    for batch, faculty in zip(labs_df['Batch'], labs_df['Faculty'])
                }
            
            # Assign the processed subject data to the final hierarchical structure
            final_consolidated_data[college][department][semester_str][division_str][subject_code] = subject_details
            
    return final_consolidated_data

def run_matrix_pipeline(matrix_file_path: str, department: str, college: str = "LDRP-ITR") -> Dict[str, Any]:
    """
    Orchestrates the entire timetable processing pipeline to generate a final
    hierarchical dictionary of consolidated timetable information, along with
    processing status and any encountered errors.

    This function calls a sequence of sub-functions to:
    1. Load and process Excel data into faculty-wise schedules.
    2. Transform faculty schedules into division-wise detailed timetables.
    3. Condense division timetables by removing time-slot and day information.
    4. Structure the condensed data into a final hierarchical dictionary.

    Args:
        matrix_file_path (str): The path to the Excel file containing timetable data.
        department (str): The name of the department (e.g., "Computer Engineering").
        college (str, optional): The name of the college. Defaults to "LDRP-ITR".

    Returns:
        Dict[str, Any]: A dictionary containing two keys:
                        - 'results': A nested dictionary containing the organized timetable data
                                     grouped by college, department, semester, division, and subject.
                                     Returns an empty dictionary if any step in the pipeline fails
                                     to produce valid data.
                        - 'status': A dictionary with processing status information:
                                    - 'success': True if the pipeline completed successfully, False otherwise.
                                    - 'message': A brief message about the processing outcome.
                                    - 'errors': A list of error/warning messages encountered during processing.
    """
    processing_errors: List[str] = []
    final_dict = {}

    # Step 1: Generate full faculty schedules
    all_faculty_schedules = process_all_timetables(matrix_file_path, errors=processing_errors)
    
    if not all_faculty_schedules:
        return {
            "results": {},
            "status": {
                "success": False,
                "message": "Failed to extract any faculty schedules. Please check the input file.",
                "errors": processing_errors
            }
        }

    # Step 2: Create division-specific timetables
    division_tables = generate_class_schedules(all_faculty_schedules, parse_errors=processing_errors)
    
    if not division_tables:
        return {
            "results": {},
            "status": {
                "success": False,
                "message": "Failed to generate division-specific timetables. This might be due to parsing errors or missing data.",
                "errors": processing_errors
            }
        }

    # Step 3: Condense division timetables
    condensed_division_tables = get_division_course_catalog(division_tables)
    
    if not condensed_division_tables:
        return {
            "results": {},
            "status": {
                "success": False,
                "message": "Failed to condense division timetables. No valid course catalog could be generated.",
                "errors": processing_errors
            }
        }

    # Step 4: Create final hierarchical dictionary
    final_dict = build_hierarchical_schedule(
        condensed_division_tables,
        department=department,
        college=college
    )
    
    if not final_dict:
        return {
            "results": {},
            "status": {
                "success": False,
                "message": "Failed to build the final hierarchical schedule. The output structure is empty.",
                "errors": processing_errors
            }
        }
    
    # If we reach here, it means the pipeline completed, even if there were some warnings/errors
    status_message = "Timetable processed successfully."
    
    if processing_errors:
        status_message = "Timetable processed with warnings/errors."

    return {
        "results": final_dict,
        "status": {
            "success": not bool(processing_errors), # True if no errors, False if any errors
            "message": status_message,
            "errors": processing_errors
        }
    }
